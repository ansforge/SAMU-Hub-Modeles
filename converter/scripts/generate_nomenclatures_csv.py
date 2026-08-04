#!/usr/bin/env python3
"""Export the nomenclatures mapping Excel file to CSV."""

from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_INPUT = (
    Path(__file__).resolve().parent.parent / "resources" / "mapping_nomenclatures.xlsx"
)


def convert(input_path: Path, output_path: Path) -> None:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        raise ValueError("Workbook contains no active worksheet")

    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    # The sheet's declared dimensions extend past the actual data into
    # trailing columns/rows that only carry styling, no values. Trim to the
    # last column that holds any real content, and drop fully-empty rows.
    last_col = max(
        (i for row in rows for i, cell in enumerate(row) if cell is not None),
        default=-1,
    )
    rows = [
        row[: last_col + 1] for row in rows if any(cell is not None for cell in row)
    ]

    with output_path.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.writer(csv_file)
        for row in rows:
            writer.writerow(["" if cell is None else cell for cell in row])


def main() -> None:
    input_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_INPUT
    if not input_path.is_file():
        print(f"Fichier introuvable : {input_path}", file=sys.stderr)
        sys.exit(1)

    output_path = input_path.parent / "export_mapping_nomenclatures.csv"
    convert(input_path, output_path)
    print(f"Export généré : {output_path}")


if __name__ == "__main__":
    main()
